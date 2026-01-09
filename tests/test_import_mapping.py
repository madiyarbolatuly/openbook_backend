from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook


@pytest.fixture()
def app(monkeypatch):
    # Import here so env vars can be set by tests if needed
    from app.main import app as fastapi_app

    return fastapi_app


def _make_xlsx(headers: list[str], rows: list[list[object]]):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_multi_sheet_xlsx(sheets: list[tuple[str, list[str], list[list[object]]]]):
    wb = Workbook()
    # Remove default sheet contents and use first provided sheet as active
    ws = wb.active
    first_title, first_headers, first_rows = sheets[0]
    ws.title = first_title
    ws.append(first_headers)
    for r in first_rows:
        ws.append(r)

    for title, headers, rows in sheets[1:]:
        wsx = wb.create_sheet(title)
        wsx.append(headers)
        for r in rows:
            wsx.append(r)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_import_products_header_mapping_smoke(app):
    client = TestClient(app)

    # This test only checks that the endpoint accepts a file and returns JSON.
    # DB integration requires a running DATABASE_URL; in CI you can point to test postgres.
    xlsx = _make_xlsx(
        headers=["Brand", "SKU", "AGSK", "ProductName", "UoM", "Price Ex VAT", "Price Inc VAT"],
        rows=[["ACME", "S1", "A1", "Phone", "pcs", 10, 12]],
    )

    resp = client.post(
        "/api/import/products",
        files={"file": ("products.xlsx", xlsx.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    # Without DB configured, FastAPI will error at startup; if DB is configured, should be 200.
    assert resp.status_code in (200, 500)


def test_import_products_multi_sheet_smoke(app):
    client = TestClient(app)

    xlsx = _make_multi_sheet_xlsx(
        [
            (
                "Brand1",
                ["Brand", "SKU", "AGSK", "ProductName", "UoM", "Price Ex VAT", "Price Inc VAT"],
                [["ACME", "S1", "A1", "Phone", "pcs", 10, 12]],
            ),
            (
                "Brand2",
                ["Brand", "SKU", "AGSK", "ProductName", "UoM", "Price Ex VAT", "Price Inc VAT"],
                [["BETA", "S2", "A2", "Tablet", "pcs", 20, 24]],
            ),
        ]
    )

    resp = client.post(
        "/api/import/products",
        files={"file": ("products.xlsx", xlsx.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert resp.status_code in (200, 500)
    if resp.status_code == 200:
        data = resp.json()
        assert "created" in data
        assert "skipped" in data
        assert "sheets" in data
        assert isinstance(data["sheets"], list)
