from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db.models import Brand, Product
from app.db.session import get_db

router = APIRouter(prefix="/api/import", tags=["import"])

# Header candidates (EN + RU) similar idea to your WinForms ImportHelper
BRAND_HEADERS = {"brand", "brandname", "бренд", "бренд наименование"}
SKU_HEADERS = {"sku", "skucode", "код", "код sku"}
AGSK_HEADERS = {"agsk", "agskcode", "agsk code", "агск"}
PRODUCT_HEADERS = {"product", "productname", "наименование", "товар"}
UOM_HEADERS = {"uom", "uom.", "uom ", "uom", "unit", "ед.изм", "ед.изм.", "единица"}
PRICE_EX_HEADERS = {"priceexvat", "price ex vat", "price exvat", "цена без ндс", "цена безндс"}
PRICE_INC_HEADERS = {"priceincvat", "price inc vat", "price incvat", "цена с ндс", "цена сндс"}


def _norm(v: object | None) -> str:
    return str(v).strip().lower() if v is not None else ""


def _build_header_map(header_row: list[object]) -> dict[str, int]:
    # maps normalized header -> 1-based index
    out: dict[str, int] = {}
    for i, val in enumerate(header_row, start=1):
        key = _norm(val)
        if key and key not in out:
            out[key] = i
    return out


def _pick_col(header_map: dict[str, int], candidates: set[str], fallback: int) -> int:
    for c in candidates:
        if c in header_map:
            return header_map[c]
    return fallback


@router.post("/products")
def import_products(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    def _to_float(v):
        if v is None or v == "":
            return None
        try:
            return float(v)
        except Exception:
            return None

    content = file.file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)

    created_total = 0
    skipped_total = 0
    sheets: list[dict[str, object]] = []

    # Some input files are organized as one brand per worksheet ("pages").
    # Process all worksheets instead of only the first one.
    for ws in wb.worksheets:
        try:
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        except StopIteration:
            # empty sheet
            sheets.append({"sheet": ws.title, "rows": 0, "created": 0, "skipped": 0, "reason": "empty"})
            continue

        header_map = _build_header_map(header)

        brand_col = _pick_col(header_map, BRAND_HEADERS, 1)
        sku_col = _pick_col(header_map, SKU_HEADERS, 2)
        agsk_col = _pick_col(header_map, AGSK_HEADERS, 3)
        product_col = _pick_col(header_map, PRODUCT_HEADERS, 4)
        uom_col = _pick_col(header_map, UOM_HEADERS, 5)
        price_ex_col = _pick_col(header_map, PRICE_EX_HEADERS, 6)
        price_inc_col = _pick_col(header_map, PRICE_INC_HEADERS, 7)

        created = 0
        skipped = 0
        rows_seen = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            rows_seen += 1
            brand_name = row[brand_col - 1] if brand_col - 1 < len(row) else None
            if brand_name is None or str(brand_name).strip() == "":
                skipped += 1
                continue

            brand_name_str = str(brand_name).strip()

            brand = db.query(Brand).filter(Brand.brand == brand_name_str).first()
            if brand is None:
                brand = Brand(brand=brand_name_str)
                db.add(brand)
                db.flush()  # get brand.id

            sku = row[sku_col - 1] if sku_col - 1 < len(row) else None
            agsk = row[agsk_col - 1] if agsk_col - 1 < len(row) else None

            # Normalize Excel values before using them in DB queries.
            # Excel might give numeric cells as int/float, but our DB columns are VARCHAR.
            sku_str = str(sku).strip() if sku is not None else None
            agsk_str = str(agsk).strip() if agsk is not None else None

            # de-dupe like .NET CreateProduct
            exists = (
                db.query(Product)
                .filter(
                    Product.brand_id == brand.id,
                    Product.sku_code == sku_str,
                    Product.agsk_code == agsk_str,
                )
                .first()
            )
            if exists is not None:
                skipped += 1
                continue

            p = Product(
                brand_id=brand.id,
                sku_code=sku_str,
                agsk_code=agsk_str,
                product=str(row[product_col - 1]).strip()
                if product_col - 1 < len(row) and row[product_col - 1] is not None
                else None,
                uom=str(row[uom_col - 1]).strip() if uom_col - 1 < len(row) and row[uom_col - 1] is not None else None,
                price_ex_vat=_to_float(row[price_ex_col - 1]) if price_ex_col - 1 < len(row) else None,
                price_inc_vat=_to_float(row[price_inc_col - 1]) if price_inc_col - 1 < len(row) else None,
            )
            db.add(p)
            created += 1

        created_total += created
        skipped_total += skipped
        sheets.append({"sheet": ws.title, "rows": rows_seen, "created": created, "skipped": skipped})

    db.commit()

    return {"created": created_total, "skipped": skipped_total, "sheets": sheets}
